import io
import openpyxl
from typing import List
from app.models import User, QueryHistory

class ExcelService:
    @staticmethod
    def generate_admin_report(users: List[User], queries: List[QueryHistory], accesses: List[dict], summary_stats: dict) -> io.BytesIO:
        """
        Generates a multi-sheet Excel report for administrators containing user tables,
        queries, district access counters, and high level summaries using pure openpyxl.
        """
        wb = openpyxl.Workbook()
        
        # Sheet 1: Users
        ws_users = wb.active
        ws_users.title = "Users"
        ws_users.append(["User ID", "Name", "Email", "Role", "Registration Date"])
        for u in users:
            ws_users.append([
                u.id,
                u.name,
                u.email,
                u.role,
                u.created_at.strftime("%Y-%m-%d %H:%M:%S") if u.created_at else ""
            ])
            
        # Sheet 2: Queries
        ws_queries = wb.create_sheet(title="Queries")
        ws_queries.append(["Query ID", "User ID", "Username", "Query", "AI Response", "District", "Date", "Time"])
        for q in queries:
            username = q.user.name if q.user else "Unknown"
            district_name = q.district.district_name if q.district else "N/A"
            ws_queries.append([
                q.id,
                q.user_id,
                username,
                q.query,
                q.response,
                district_name,
                q.created_at.strftime("%Y-%m-%d") if q.created_at else "",
                q.created_at.strftime("%H:%M:%S") if q.created_at else ""
            ])
            
        # Sheet 3: District Access
        ws_access = wb.create_sheet(title="District Access")
        ws_access.append(["District", "Total Views", "Unique Users", "Last Accessed"])
        for a in accesses:
            last_acc = a.get("last_accessed")
            last_acc_str = last_acc.strftime("%Y-%m-%d %H:%M:%S") if last_acc else "N/A"
            ws_access.append([
                a.get("district_name"),
                a.get("total_views", 0),
                a.get("unique_users", 0),
                last_acc_str
            ])
            
        # Sheet 4: Summary
        ws_summary = wb.create_sheet(title="Summary")
        ws_summary.append(["Metric", "Value"])
        ws_summary.append(["Total Users", summary_stats.get("total_users", 0)])
        ws_summary.append(["Total Queries", summary_stats.get("total_queries", 0)])
        ws_summary.append(["Total Districts", summary_stats.get("total_districts", 0)])
        ws_summary.append(["Most Accessed District", summary_stats.get("most_viewed_district", "None")])
        ws_summary.append(["Average Queries Per User", round(summary_stats.get("avg_queries_per_user", 0.0), 2)])
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
